import os
import sys

# 添加项目根目录到Python路径
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
if project_root not in sys.path:
    sys.path.append(project_root)


from configs import config as config


import uuid
import json
import pymysql

from apps.app_full import *
from models.langchain_models import embedding_bge
from sklearn.metrics.pairwise import cosine_similarity

from openpyxl import load_workbook, Workbook


class Dataset:
    def __init__(self, path='/data/liyiru/WiSearch/eval/Query.xlsx'):
        wb = load_workbook(path)
        sheet = wb['结构化查询']

        self.query = []
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                self.query.append(row[0])

        '''self.result = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                self.result.append(row[0])'''

        self.sql = []
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                self.sql.append(row[0])

def eval(dataset, pipeline, conn):

    cursor = conn.cursor()
    
    result = {}

    std_query = dataset.query
    # std_result = dataset.result
    std_sql = dataset.sql

    correct = 0
    avg_similarity = 0

    total = 0
    poi = 0
    time_mask = 0
    time_process = 0
    schema_link = 0
    generate = 0


    for i in range(len(std_query)):

        # sql
        sql, time_dict = pipeline(std_query[i])
        total += time_dict['total']
        poi += time_dict['poi_time']
        time_mask += time_dict['time_mask_time']
        time_process += time_dict['time_process_time']
        schema_link += time_dict['schema_link_time']
        generate += time_dict['generate_time']
        
        emb_pipeline = embedding_bge(sql).reshape(1, -1)
        emb_std = embedding_bge(std_sql[i])
        # sql相似度
        sql_similarity = cosine_similarity(emb_pipeline, emb_std)[0][0]
        avg_similarity += sql_similarity

        cursor.execute(sql)
        print("SQL:", sql)
        result_pipeline = cursor.fetchall()
        print("STD:", std_sql[i])
        cursor.execute(std_sql[i])
        result_std = cursor.fetchall()

        # 结果
        if result_pipeline == result_std:
            correct += 1
            correctness = "CORRECT"
        else:
            correctness = "INCORRECT"

        result_list = []
        std_list = []
        print(result_pipeline)
        for item in result_pipeline:
            print("ITEM:", item)
            try:
                result_list.append(f"{item[0]} {item[1]}")
            except:
                result_list.append(f"{item[0]}")
        final_result = '\n'.join(result_list)

        print(result_std)
        for item in result_std:
            print("STD ITEM:", item)
            try:
                std_list.append(f"{item[0]} {item[1]}")
            except:
                std_list.append(f"{item[0]}")
        final_std = '\n'.join(std_list)

        result[f'{i}'] = {"QUERY": std_query[i], "CORRECTNESS": correctness, "RESULT": final_result, "RESULT_STD": final_std, "SQL SIMILARITY": sql_similarity, "SQL": sql, "SQL_STD": std_sql[i]}

    total_number = len(std_query)
    result['avg_time'] = {'total': f'总耗时: {round(total / total_number, 2)}', 'poi_time': f'关键点抽取: {round(poi / total_number, 2)}', 'time_mask_time': f'时间掩码: {round(time_mask / total_number, 2)}', 'time_process_time': f'时间处理: {round(time_process / total_number, 2)}', 'schema_link_time': f'元数据关联: {round(schema_link / total_number, 2)}', 'generate_time': f'生成: {round(generate / total_number, 2)}'}

    accuracy = round(correct / len(std_query) * 100, 2)
    similarity = round(avg_similarity / len(std_query) * 100, 2)



    print("RESULT:", result)


    return accuracy, similarity, result




if __name__ == "__main__":
    dataset = Dataset(path='/data/liyiru/WiSearch/eval/Query.xlsx')
    pipeline_type = 'app_full'
    # print(dataset.query)
    # print(dataset.result)
    # print(dataset.sql)

    db_conn = pymysql.connect(
    host='172.31.24.111',
    user='root',
    password='liucd123',
    database='12345',
    port=3307,
    charset='utf8mb4',
    )

    # cursor = db_conn.cursor()

    # cursor.execute("SELECT `诉求区域`, COUNT(*) AS 总数 FROM shanghai_ad_time WHERE `工单类型`='投诉举报类' AND `四级分类`='网上购物' AND `工单生成时间`> DATE_SUB(CURDATE(), INTERVAL 6 MONTH) GROUP BY `诉求区域` ORDER BY 总数 DESC")
    emb_1 = embedding_bge("投诉事件").reshape(1, -1)
    emb_2 = embedding_bge("投诉举报类").reshape(1, -1)

    result = cosine_similarity(emb_1, emb_2)[0][0]
    print(result)

    accuracy, similarity, result = eval(dataset, main, db_conn)
    # accuracy = 50
    # result = {"1": {"QUERY": "TEST1", "CORRECTNESS": "TEST2", "RESULT": "TEST3", "RESULT_STD": "TEST4", "SQL SIMILARITY": "TEST5", "SQL": "TEST6", "SQL_STD": "TEST7"}, "2": {"QUERY": "TEST8", "CORRECTNESS": "TEST9", "RESULT": "TEST10", "RESULT_STD": "TEST11", "SQL SIMILARITY": "TEST12", "SQL": "TEST13", "SQL_STD": "TEST14"}}
    # 表头
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'模型：{pipeline_type}'
    ws['B1'] = f'数据集：12345-shanghai_ad_time'
    ws['C1'] = f'准确率：{accuracy}%'
    ws['D1'] = f'SQL相似度: {similarity}%'
    ws['A2'] = '用户提问'
    ws['B2'] = '结果是否正确'
    ws['C2'] = '模型生成结果'
    ws['D2'] = '标准结果'
    ws['E2'] = 'SQL相似度'
    ws['F2'] = '模型生成SQL'
    ws['G2'] = '标准SQL'

    # 内容
    for key, value in result.items():
        ws.append(list(value.values()))
    if os.path.exists(f'./eval/{pipeline_type}.xlsx'):
        os.remove(f'./eval/{pipeline_type}.xlsx')
    wb.save(f'./eval/{pipeline_type}.xlsx')


