from pathlib import Path
import sys 
import os 
# 添加项目根目录到Python路径
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
if project_root not in sys.path:
    sys.path.append(project_root)


from models.langchain_models import llm_qwen_7B, embedding_bge
from utils.util import *

from langchain.prompts import PromptTemplate

from configs import config as config
from offline_initial_file import init
from langchain_community.utilities import SQLDatabase

from models.create_chain import *

from sql_processing.text2sql import schema_linking
from schema_linking.chain_link import chain_link
from schema_linking.case_retrieval import retrieve_cases

from chathistory.history import ChatHistory, history_preprocess

import uuid
import json
import time
from datetime import datetime
import pymysql
import ast



InitChatHistory = ChatHistory()

# query = 'Dane的项目有哪些'

def main(query):
    '''if session_id == '':
        session_id = str(uuid.uuid4())
    message_id = str(uuid.uuid4())'''
    # history = InitChatHistory.retrieve_history(session_id)
    '''if history != []:
        history_string = history_preprocess(history)
        print("HISTORY:", history_string)
        rewrite_query, status = rewrite(query, history_string)
        if status != '需要':
            InitChatHistory.clear_history(session_id)
    else:
        rewrite_query = query

    InitChatHistory.update_history(session_id, message_id, f'Human: {rewrite_query}')'''
    
    # 假设性SQL
    # st = time.time()
    # assumption = assumption_chain.invoke({"query":query})
    # print("ASSUMPTION:", assumption, query)
    # ft = time.time()

    # print(ft - st)

    st_main = time.time()

    # SQL元数据
    with open("metadata/metadata_12345_shanghai_ad_time.txt", 'r', encoding='utf-8') as f:
        content = f.read()
    # SQL枚举值
    with open('metadata/related_values_shanghai_ad_time.json', 'r', encoding='utf-8') as f:
            related_values = json.loads(f.read())

    # 1. 时间mask
    st = time.time()
    time_mask = time_mask_chain.invoke({"query": query})
    print("TIME:", time_mask, query)
    time_mask_query = time_mask['masked_query']
    ft = time.time()
    time_mask_time = ft - st
    print(ft - st)

    # 2. 时间处理
    st = time.time()
    time_process = time_process_chain.invoke({"query": time_mask['time_mask']})
    print("TIME P:", time_process, query)
    time_p = time_process['time']
    ft = time.time()
    print(ft - st)
    time_process_time = ft - st

    # 3. 询问意图理解
    st = time.time()
    intention_process = intention_mask_chain.invoke({"query": query})
    print("INTENTION:", intention_process, query)
    intention = intention_process['intention']
    key_point = intention_process['key_point']
    ft = time.time()
    print(ft - st)
    intention_process_time = ft - st

    # 4. 关键点识别
    st = time.time()
    poi = poi_mask_chain.invoke({"query": time_mask_query, "metadata":content, "values":related_values, "intention": intention_process['intention'], "key_point": intention_process['key_point']})
    print("POI:", poi)
    if type(poi['key_information']) == str:
        key_information_list = ast.literal_eval(poi['key_information'])
    else:
        key_information_list = poi['key_information']
    print("KEY INFORMATION:", key_information_list)
    ft = time.time()
    print(ft - st)
    poi_process_time = ft - st

    # 5. schema link
    st = time.time()
    try:
        link_result = f"工单生成时间: {time_p}"
    except:
        link_result = ''

    # linking = schema_linking_chain.invoke({"query": query, "schema": content, "samples": samples})
    for i in range(len(key_information_list)):
        mask = key_information_list[i]
        
        linking = retrieve_cases_full(query=query, metadata=content, related_values=related_values, key=mask)
        print("SL:", linking, mask)
        link_result += f' {linking[0]}: {linking[1]}'

    ft = time.time()
    print(ft - st)
    schema_link_time = ft - st

    # 6. 生成
    st = time.time()
    generate = sql_gen_mask_chain.invoke({"query": query, "table": content, "time": link_result, "intention": key_point + intention})
    print(generate)
    ft = time.time()
    print(ft - st)
    generate_time = ft - st

    # 7. 错误重写，没有关联数据库则不需要这步
    st = time.time()

    db_conn = pymysql.connect(
        host='172.31.24.111',
        user='root',
        password='liucd123',
        database='12345',
        port=3307,
        charset='utf8mb4'
        )
    
    cursor = db_conn.cursor()

    is_runnable = False
    while not is_runnable:
        try:
            cursor.execute(generate['SQL'])
            is_runnable = True
        except Exception as e:
            generate = error_rewrite_chain.invoke({"query": query, "table": content, "time": link_result, "intention": intention, "generated_sql": generate['SQL'], "error_message": e})
    
    print(generate)
    ft = time.time()
    print(ft - st)

    # 8. 处理输出结果
    error_time = ft - st


    ft_main = time.time()
    total_time = ft_main - st_main

    return generate['SQL'], {'total': total_time, 'poi_process_time': poi_process_time, 'time_mask_time': time_mask_time, 'time_process_time': time_process_time, 'intention_process_time': intention_process_time, 'schema_link_time': schema_link_time, 'generate_time': generate_time, 'error_time': error_time}

    

    # schema关联
    # linking = schema_linking_chain.invoke({"query": query, "schema": content, "samples": samples})

    # generate = sql_gen_mask_chain.invoke({"query": time_mask_query, "table": content, "time": time_result})

    # print("GENERATE:", generate, query)
    

    # 选择
    

    # return gen_result_1

from flask import Flask, request, jsonify, Response
import requests

app = Flask(__name__)
@app.route('/gen', methods=['POST'])
def nl2sql():
    input_data = request.get_data()
    query = json.loads(input_data)['query']
    sql, time = main(query=query)
    return jsonify({"SQL": sql})


if __name__ == '__main__':
    '''session_id = str(uuid.uuid4())
    message_id_1 = str(uuid.uuid4())
    message_id_2 = str(uuid.uuid4())'''
    # InitChatHistory.update_history(session_id, message_id_1, 'HUMAN: 2024年的一审判决案件中，有多少案件的嫌疑人是男性')
    # InitChatHistory.update_history(session_id, message_id_2, 'AI: 2024年的一审判决案件中，嫌疑人是男性的案件有15起')
    '''retrieved = InitChatHistory.retrieve_history(session_id)
    print(retrieved)'''
    # query = '近1个月虹口区发生了多少起城市管理类的投诉事件'
    # main("今年上半年企业服务工单里工单类型分别有那些，占比如何")
    # main("近1个月虹口区发生了多少起城市管理类的投诉事件")


    '''query = '小区居改非的工单中，投诉的比例是多少'
    with open("metadata/metadata_12345_shanghai_ad_time.txt", 'r', encoding='utf-8') as f:
        content = f.read()

    link_result = '四级分类: 居改非 工单类型: 投诉举报类'
    generate = sql_gen_mask_chain.invoke({"query": query, "table": content, "time": link_result})

    print(generate)'''

    # main("近半年医保相关的咨询事件有多少起")
    # app.run(host='0.0.0.0', debug=False, port=33072)


    from openpyxl import load_workbook, Workbook

    wb = load_workbook('/data/liyiru/WiSearch/eval/Query.xlsx')
    sheet = wb['结构化查询']

    query = []
    intentions = []
    key_points = []
    key_informations = []
    # masked = []
    # masks = []
    # mask_map = []

    with open("metadata/metadata_12345_shanghai_ad_time.txt", 'r', encoding='utf-8') as f:
        content = f.read()
        print("METADATA:", content)

    with open('metadata/related_values_shanghai_ad_time.json', 'r', encoding='utf-8') as f:
        related_values = json.loads(f.read())
        print("VALUES:", related_values)

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=sheet.max_row, values_only=True):
        if row[0] is not None:
            query.append(row[0])
            print("QUERY:", row[0])
            time_mask = time_mask_chain.invoke({"query": row[0]})
            print("TIME:", time_mask)
            time_mask_query = time_mask['masked_query']
            intention_process = intention_mask_chain.invoke({"query": time_mask_query})
            print("INTENTION:", intention_process)
            intentions.append(intention_process['intention'])
            key_points.append(intention_process['key_point'])
            poi = poi_mask_chain.invoke({"query": time_mask_query, "metadata":content, "values":related_values, "intention": intention_process['intention'], "key_point": intention_process['key_point']})
            print("POI:", poi)
            if type(poi['key_information']) == str:
                key_information_list = ast.literal_eval(poi['key_information'])
            else:
                key_information_list = poi['key_information']
            print("KEY INFORMATION:", key_information_list)
            key_informations.append(','.join(key_information_list))
            # masked.append(poi['masked_query'])
            # masks.append(','.join(poi['masks'][:10]))
            # mask_map.append(','.join(poi['mask_map'][:10]))
            # print("POI:", poi)

            for i in range(len(key_information_list)):
                mask = key_information_list[i]
                
                linking = retrieve_cases_full(query=row[0], metadata=content, related_values=related_values, key=mask)
                print("SL:", linking, mask)

    new_wb = Workbook()
    ws = new_wb.active
    ws['A1'] = 'query'
    ws['B1'] = 'intention'
    ws['C1'] = 'key_point'
    ws['D1'] = 'key_information'

    for i in range(len(query)):
        print("SAVING:", [query[i],intentions[i],key_points[i],key_informations[i]])
        ws.append([query[i],intentions[i],key_points[i],key_informations[i]])
    if os.path.exists(f'./eval/poi2.xlsx'):
        os.remove(f'./eval/poi2.xlsx')
    new_wb.save(f'./eval/poi2.xlsx')